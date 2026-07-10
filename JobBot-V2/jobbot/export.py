"""Spreadsheet exports: formatted Excel workbook and plain CSV."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Score", "score", 8),
    ("Status", "status", 13),
    ("Title", "title", 38),
    ("Company", "company", 24),
    ("Location", "location", 22),
    ("Remote", "is_remote", 9),
    ("Salary Min", "salary_min", 12),
    ("Salary Max", "salary_max", 12),
    ("Posted", "date_posted", 12),
    ("Source", "source", 10),
    ("Fit Summary", "fit_summary", 50),
    ("Strengths", "strengths", 45),
    ("Gaps", "gaps", 45),
    ("Notes", "notes", 30),
    ("URL", "url", 45),
    ("Job ID", "id", 14),
]


def _cell_value(job: dict, key: str):
    value = job.get(key)
    if key == "is_remote":
        return "Yes" if value else "No"
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return value


def timestamped(prefix: str, ext: str) -> str:
    return f"{prefix}_{datetime.now():%Y%m%d_%H%M%S}.{ext}"


def export_csv(jobs: list[dict], out_path: Path) -> Path:
    with out_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow([header for header, _, _ in COLUMNS])
        for job in jobs:
            writer.writerow([_cell_value(job, key) for _, key, _ in COLUMNS])
    return out_path


def export_xlsx(jobs: list[dict], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    hi_fill = PatternFill("solid", fgColor="DCFCE7")   # score >= 80
    mid_fill = PatternFill("solid", fgColor="FEF3C7")  # score >= 65
    wrap = Alignment(wrap_text=True, vertical="top")

    for row_idx, job in enumerate(jobs, 2):
        for col, (_, key, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col, value=_cell_value(job, key))
            if key in ("fit_summary", "strengths", "gaps", "notes"):
                cell.alignment = wrap
            if key == "url" and job.get("url"):
                cell.hyperlink = job["url"]
                cell.font = Font(color="2563EB", underline="single")
        score = job.get("score") or 0
        score_cell = ws.cell(row=row_idx, column=1)
        score_cell.font = Font(bold=True)
        if score >= 80:
            score_cell.fill = hi_fill
        elif score >= 65:
            score_cell.fill = mid_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(jobs) + 1, 2)}"
    wb.save(out_path)
    return out_path
